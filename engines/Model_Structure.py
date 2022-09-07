import math, os
from engines.utils import metrics, extractEntity
import tensorflow as tf
from bert import modeling
from bert import tokenization
import pandas as pd
import time,re
import xlwt,xlrd
from stanfordcorenlp import StanfordCoreNLP
from xlrd import open_workbook
from collections import Counter
import itertools
import numpy as np
import openpyxl
import logging
import pickle

tf.logging.set_verbosity(tf.logging.ERROR)
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

def collect_final_step_lstm(lstm_rep, lens):
    lens = tf.maximum(lens, tf.zeros_like(lens, dtype=tf.int32)) # [batch,]
    idxs = tf.range(0, limit=tf.shape(lens)[0]) # [batch,]
    indices = tf.stack((idxs,lens,), axis=1) # [batch_size, 2]
    return tf.gather_nd(lstm_rep, indices, name='lstm-forward-last')

class BertEncoder(object):
    def __init__(self, model_root, bert_config_file, init_checkpoint, vocab_file, max_sequence_len, embedding_batch,
                 sen2id_path, embedding_matrix_path, vec_dim):
        self.model_root = model_root
        self.bert_config = modeling.BertConfig.from_json_file(bert_config_file)
        self.init_checkpoint = init_checkpoint
        self.vocab_file = vocab_file
        self.max_sequence_len = max_sequence_len
        self.embedding_batch = embedding_batch
        self.vec_dim = vec_dim
        self.embedding_matrix_path = embedding_matrix_path
        self.sen2id = load_vocabulary_from_file(sen2id_path)
        self.input_ids = tf.placeholder(tf.int32, shape=[None, None], name='input_ids')  # [batch_size, sentence_length]
        self.input_mask = tf.placeholder(tf.int32, shape=[None, None],
                                         name='input_masks')  # [batch_size, sentence_length]
        self.segment_ids = tf.placeholder(tf.int32, shape=[None, None],
                                          name='segment_ids')  # [batch_size, sentence_length]
        self.bert_model = modeling.BertModel(
            config=self.bert_config,
            is_training=True,
            input_ids=self.input_ids,
            input_mask=self.input_mask,
            token_type_ids=self.segment_ids,
            use_one_hot_embeddings=False
        )
        tvars = tf.trainable_variables()
        (self.assignment_map, _) = modeling.get_assignment_map_from_checkpoint(tvars, self.init_checkpoint)
        tf.train.init_from_checkpoint(self.init_checkpoint, self.assignment_map)
        self.encoder_out = self.bert_model.get_sequence_output()

    def encode_by_model(self, corpus, save_path=None):
        with tf.Session() as sess:
            sess.run(tf.global_variables_initializer())
            if isinstance(corpus, str):
                word_ids, word_mask, word_segment_ids = prepare_input(corpus, self.vocab_file, self.max_sequence_len)
                feed_dict = {self.input_ids: np.array(word_ids), self.input_mask: np.array(word_mask),
                             self.segment_ids: np.array(word_segment_ids)}
                corpus_vector = sess.run(self.encoder_out, feed_dict)
                corpus_vector = extract_embedding(corpus_vector)
                print('last shape:{}'.format(corpus_vector.shape))
            else:
                assert isinstance(corpus, list)
                id2sen = reverse_dict(self.sen2id)
                assert len(self.sen2id) == len(id2sen)
                nb_batch, batch_indexs = (len(corpus), self.embedding_batch)
                bert_embedding_matrix = np.zeros([len(corpus), self.max_sequence_len, self.vec_dim])
                for batch_index in range(nb_batch):
                    if batch_index % 10 == 0:
                        logging.info("'current bert embedding_batch index:{}'.format(batch_index)")
                    cur_batch = batch_indexs[batch_index]
                    cur_corpus = (cur_batch, id2sen)
                    cur_word_ids, cur_word_mask, cur_word_segment_ids = prepare_input(cur_corpus, self.vocab_file,
                                                                                      self.max_sequence_len)
                    feed_dict = {self.input_ids: cur_word_ids, self.input_mask: cur_word_mask,
                                 self.segment_ids: cur_word_segment_ids}
                    cur_corpus_vector = sess.run(self.encoder_out, feed_dict)
                    assert len(cur_corpus_vector.shape) == 3
                    for sen_index in range(cur_corpus_vector.shape[0]):
                        bert_embedding_matrix[self.sen2id[cur_corpus[sen_index]], :, :] = turn_embedding(
                            cur_corpus_vector[sen_index, :, :], self.max_sequence_len, self.vec_dim,
                            len(cur_corpus[sen_index]))
                if save_path:
                    with open(save_path, 'wb') as outfile:
                        pickle.dump(bert_embedding_matrix, outfile)
                print('last shape:{}'.format(bert_embedding_matrix.shape))
                return bert_embedding_matrix

def turn_embedding(corpus_vector, max_sequence_len, vec_dim, sen_len):
    """
     distill vector from model's output
    """
    assert len(corpus_vector.shape) == 2

    # corpus_vector[0, :] = zero_vector
    end_corpus_vector = np.zeros([max_sequence_len, vec_dim])
    corpus_vector = corpus_vector[:sen_len + 2, :]
    corpus_vector = extract_embedding(corpus_vector)
    end_corpus_vector[:corpus_vector.shape[0], :] = corpus_vector
    return end_corpus_vector


def extract_embedding(corpus_vector):
    # wipe off representation for [cls] and [sep]
    assert len(corpus_vector.shape) == 2 or len(corpus_vector.shape) == 3
    if len(corpus_vector.shape) == 3:
        return corpus_vector[:, 1: corpus_vector.shape[1] - 1, :]
    return corpus_vector[1: corpus_vector.shape[0] - 1, :]


def load_vocabulary_from_file(voc_path):
    with open(voc_path, 'rb') as infile:
        voc = pickle.load(infile)
    return voc


def prepare_input(corpus, vocab_file, max_sequence_len):
    """
    :return: id,  mask, type for model input
    """
    token = tokenization.CharTokenizer(vocab_file)
    if isinstance(corpus, list):
        word_ids = []
        word_mask = []
        word_segment_ids = []

        for sentence in corpus:
            sen_token = []
            sen_segment_id = []
            sen_token.append("[CLS]")
            sen_segment_id.append(0)
            sen_split_tokens = token.tokenize(sentence)
            for token_ in sen_split_tokens:
                sen_token.append(token_)
                sen_segment_id.append(0)
            sen_token.append("[SEP]")
            sen_segment_id.append(0)
            sen_word_ids = token.convert_tokens_to_ids(sen_token)
            sen_word_mask = [1] * len(sen_word_ids)

            while len(sen_word_ids) < max_sequence_len:
                sen_word_ids.append(0)
                sen_word_mask.append(0)
                sen_segment_id.append(0)

            assert len(sen_word_ids) == max_sequence_len
            assert len(sen_word_mask) == max_sequence_len
            assert len(sen_segment_id) == max_sequence_len

            word_ids.append(sen_word_ids)
            word_mask.append(sen_word_mask)
            word_segment_ids.append(sen_segment_id)
        return np.array(word_ids), np.array(word_mask), np.array(word_segment_ids)

    else:
        assert isinstance(corpus, str)
        split_tokens = token.tokenize(corpus)
        word_token = []
        word_token.append("[CLS]")
        for token_ in split_tokens:
            word_token.append(token_)
        word_token.append("[SEP]")
        word_ids = token.convert_tokens_to_ids(word_token)
        word_mask = [1] * len(word_ids)
        word_segment_ids = [0] * len(word_ids)
        return np.array([word_ids]), np.array([word_mask]), np.array([word_segment_ids])

def reverse_dict(dict_):
    assert isinstance(dict_, dict)
    return {v: k for k, v in dict_.items()}
    
def Sentence_to_Paper(input_path_M, input_path_E,papaer_unit_excel):
    def load_file(path):
        workbook2 = open_workbook(path)
        sheet2 = workbook2.sheet_by_name('Sheet')
        paper_id = sheet2.col_values(0)[1:]
        title = sheet2.col_values(1)[1:]
        meeting = sheet2.col_values(2)[1:]
        year = sheet2.col_values(3)[1:]
        para_location = sheet2.col_values(4)[1:]
        sentences = sheet2.col_values(5)[1:]
        method_entities = sheet2.col_values(9)[1:]
        experiment_entities = sheet2.col_values(10)[1:]
        metrics_entities = sheet2.col_values(11)[1:]
        return paper_id, title, meeting, year, para_location, sentences, method_entities, experiment_entities, metrics_entities

    def count_number(lst):
        dict_lst = dict.fromkeys(lst, 0)
        for x in lst:
            dict_lst[x] += 1
        dict_lst_sort = sorted(dict_lst.items(), key=lambda x: x[1], reverse=True)
        return dict(dict_lst_sort)

    def get_calss(flag_number):
        paper_number = []
        for c_fn in flag_number:
            if c_fn not in paper_number:
                paper_number.append(c_fn)
        paper_split_flag = []
        for row_pn in paper_number:
            temp = []
            for i_n, row_n in enumerate(flag_number):
                if row_n == row_pn:
                    temp.append(i_n)
            paper_split_flag.append(temp)
        return paper_split_flag

    def get_commen_paper(paper_split_flag, body):
        paper_split_content = []
        for row_psf in paper_split_flag:
            temp1 = []
            for r_row_psf in row_psf:
                temp1.append(body[r_row_psf])
            paper_split_content.append(temp1)
        return paper_split_content

    def get_entities(method_entities_split):
        method_entities_split_notnan_all_paper = []
        for one_paper_entities in method_entities_split:
            method_entities_split_notnan_one_paper = []
            for one_entity in one_paper_entities:
                if type(
                        one_entity) == str and one_entity not in method_entities_split_notnan_one_paper and one_entity != "":
                    method_entities_split_notnan_one_paper.append(one_entity)
            method_entities_split_notnan_all_paper.append(method_entities_split_notnan_one_paper)
        return method_entities_split_notnan_all_paper

    def get_paper_unit_content(paper_split_flag, paper_id, title, meeting, year, method_entities, dateset_entities,
                               metrics_entities):
        paper_id_split = [c_paper_id[0] for c_paper_id in get_commen_paper(paper_split_flag, paper_id)]
        title_split = [c_title[0] for c_title in get_commen_paper(paper_split_flag, title)]
        meeting_split = [c_meeting[0] for c_meeting in get_commen_paper(paper_split_flag, meeting)]
        year_split = [c_year[0] for c_year in get_commen_paper(paper_split_flag, year)]
        method_entities_split = get_commen_paper(paper_split_flag, method_entities)
        method_entities_split_notnan_all_paper = get_entities(method_entities_split)
        dataset_entities_split = get_commen_paper(paper_split_flag, dateset_entities)
        dataset_entities_split_notnan_all_paper = get_entities(dataset_entities_split)
        metrics_entities_split = get_commen_paper(paper_split_flag, metrics_entities)
        metrics_entities_split_notnan_all_paper = get_entities(metrics_entities_split)
        return paper_id_split, title_split, meeting_split, year_split, method_entities_split_notnan_all_paper, dataset_entities_split_notnan_all_paper, metrics_entities_split_notnan_all_paper

    def get_frequency(method_entities_split_notnan_all_paper):
        allpapermethods = []
        for one_paper_methods in method_entities_split_notnan_all_paper:
            temp = []
            if len(one_paper_methods) > 0:
                for one_sentence_methods in one_paper_methods:
                    one_sentence_methods_split = one_sentence_methods.split(";")
                    temp.extend(one_sentence_methods_split)
            frequency = count_number(temp)
            allpapermethods.append(frequency)
        return allpapermethods

    def getallpaperMDM(input_path):
        paper_id, title, meeting, year, para_location, sentences, method_entities, dateset_entities, metrics_entities = load_file(
            input_path)
        paper_split_flag = get_calss(paper_id)
        paper_id_split, title_split, meeting_split, year_split, method_entities_split_notnan_all_paper, dataset_entities_split_notnan_all_paper, metrics_entities_split_notnan_all_paper = get_paper_unit_content(
            paper_split_flag, paper_id, title, meeting, year, method_entities, dateset_entities, metrics_entities)
        allpapermethods = get_frequency(method_entities_split_notnan_all_paper)
        allpaperdatasets = get_frequency(dataset_entities_split_notnan_all_paper)
        allpapermetrics = get_frequency(metrics_entities_split_notnan_all_paper)
        return paper_id_split, title_split, meeting_split, year_split, allpapermethods, allpaperdatasets, allpapermetrics

    def get_df_M(paper_id_split_M, title_split_M, meeting_split_M, year_split_M, allpapermethods_M, allpaperdatasets_M,
                 allpapermetrics_M):
        df_M = pd.DataFrame({'paper_id': paper_id_split_M,
                             'title': title_split_M,
                             'meeting': meeting_split_M,
                             'year': year_split_M,
                             'methods_M': allpapermethods_M,
                             'datasets_M': allpaperdatasets_M,
                             'metrics_M': allpapermetrics_M},
                            index=paper_id_split_M)
        return df_M

    def get_df_E(paper_id_split_M, title_split_M, meeting_split_M, year_split_M, allpapermethods_M, allpaperdatasets_M,
                 allpapermetrics_M):
        df_E = pd.DataFrame({'paper_id_E': paper_id_split_M,
                             'title_E': title_split_M,
                             'meeting_E': meeting_split_M,
                             'year_E': year_split_M,
                             'methods_E': allpapermethods_M,
                             'datasets_E': allpaperdatasets_M,
                             'metrics_E': allpapermetrics_M},
                            index=paper_id_split_M)
        return df_E

    def comine_every_dict(input_x, input_y):
        if type(input_x) == dict and type(input_y) == dict:
            input_X, input_Y = Counter(input_x), Counter(input_y)
            combine_dict_result = dict(input_X + input_Y)
        elif type(input_x) == dict and type(input_y) != dict:
            combine_dict_result = input_x
        elif type(input_x) != dict and type(input_y) == dict:
            combine_dict_result = input_y
        else:
            combine_dict_result = 0
        combine_dict_result_lst = sorted(combine_dict_result.items(), key=lambda x: x[1], reverse=True)
        return dict(combine_dict_result_lst)

    def get_add_ME(df_ME, colName_M, colName_E, colfinalName):
        methods_M = df_ME[colName_M].values
        methods_E = df_ME[colName_E].values
        combine_methods_ME = []
        for i in range(len(methods_M)):
            every_row = comine_every_dict(methods_M[i], methods_E[i])
            combine_methods_ME.append(every_row)

        df_methods_ME = pd.DataFrame({colfinalName: combine_methods_ME},
                                     index=df_ME.index)
        return df_methods_ME

    def get_value(dict_ME, dict_M, dict_E):
        for key in dict_ME.keys():
            try:
                key_de_value_M = dict_M[key]
            except:
                key_de_value_M = 0
            try:
                key_de_value_E = dict_E[key]
            except:
                key_de_value_E = 0
            dict_ME[key] = [dict_ME[key], key_de_value_M, key_de_value_E]
        return dict_ME

    def get_df_mdm_all(df_methods_ME, colName1, df_ME, cloName2, colName3, preserve_colName):
        lst_methods_ME = df_methods_ME[colName1].values.tolist()
        lst_methods_M = df_ME[cloName2].values.tolist()
        lst_methods_E = df_ME[colName3].values.tolist()
        for id_dict in range(len(lst_methods_ME)):
            get_value(lst_methods_ME[id_dict], lst_methods_M[id_dict], lst_methods_E[id_dict])
        df_methods = pd.DataFrame({preserve_colName: lst_methods_ME}, index=df_ME.index)
        return df_methods


    paper_id_split_M, title_split_M, meeting_split_M, year_split_M, allpapermethods_M, allpaperdatasets_M, allpapermetrics_M = getallpaperMDM(
        input_path_M)
    df_M = get_df_M(paper_id_split_M, title_split_M, meeting_split_M, year_split_M, allpapermethods_M,
                    allpaperdatasets_M, allpapermetrics_M)
    paper_id_split_E, title_split_E, meeting_split_E, year_split_E, allpapermethods_E, allpaperdatasets_E, allpapermetrics_E = getallpaperMDM(
        input_path_E)
    df_E = get_df_E(paper_id_split_E, title_split_E, meeting_split_E, year_split_E, allpapermethods_E,
                    allpaperdatasets_E, allpapermetrics_E)
    frames = [df_M, df_E]
    df_ME = pd.concat(frames, axis=1)
    df_ME["paper_id"] = df_ME.index
    df_methods_ME = get_add_ME(df_ME, "methods_M", "methods_E", 'methods_ME')
    df_datasets_ME = get_add_ME(df_ME, "datasets_M", "datasets_E", "datasets_ME")
    df_metrics_ME = get_add_ME(df_ME, "metrics_M", "metrics_E", "metrics_ME")
    df_methods = get_df_mdm_all(df_methods_ME, "methods_ME", df_ME, "methods_M", "methods_E", 'methods')
    df_datasets = get_df_mdm_all(df_datasets_ME, "datasets_ME", df_ME, "datasets_M", "datasets_E", 'datasets')
    df_metrics = get_df_mdm_all(df_metrics_ME, "metrics_ME", df_ME, "metrics_M", "metrics_E", 'metrics')
    frames_2 = [df_ME, df_methods, df_datasets, df_metrics]
    df_ME_mdm = pd.concat(frames_2, axis=1)
    df_final_result_origorder = df_ME_mdm.drop(
        columns=["methods_M", "datasets_M", "metrics_M", 'paper_id_E', 'title_E', 'meeting_E', 'year_E',
                 "methods_E", "datasets_E", "metrics_E"])
    df_final_result_origorder.to_excel(papaer_unit_excel)



class Model_Structure(object):
    def __init__(self, configs, logger, dataManager):
        os.environ['CUDA_VISIBLE_DEVICES'] = configs.CUDA_VISIBLE_DEVICES

        self.configs = configs
        self.logger = logger
        self.logdir = configs.log_dir
        self.measuring_metrics = configs.measuring_metrics
        self.dataManager = dataManager

        if configs.mode == "train":
            self.is_training = True
        else:
            self.is_training = False

        self.checkpoint_name = configs.checkpoint_name
        self.checkpoints_dir = configs.checkpoints_dir
        self.output_test_file = configs.datasets_fold + "/" + configs.output_test_file
        self.is_output_sentence_entity = configs.is_output_sentence_entity
        self.output_sentence_entity_file = configs.datasets_fold + "/" + configs.output_sentence_entity_file
        self.is_real_output_sentence_entity = configs.is_real_output_sentence_entity
        self.real_output_sentence_entity_file = configs.datasets_fold + "/" + configs.real_output_sentence_entity_file
        self.is_compare_output_sentence_entity = configs.is_compare_output_sentence_entity
        self.compare_output_sentence_entity_file = configs.datasets_fold + "/" + configs.compare_output_sentence_entity_file
        self.is_real_test = configs.is_real_test
        self.output_excel = configs.output_excel
        self.raw_real_sens_file = configs.raw_real_sens_file
        self.output_methods_file = configs.output_methods_excel
        self.output_experiment_file = configs.output_experiment_excel
        self.papaer_unit_excel_file = configs.output_paper_unit_excel

        self.biderectional = configs.biderectional
        self.cell_type = configs.cell_type
        self.num_layers = configs.encoder_layers

        self.is_crf = configs.use_crf

        self.learning_rate = configs.learning_rate
        self.dropout_rate = configs.dropout
        self.batch_size = configs.batch_size

        self.emb_dim = configs.embedding_dim
        self.hidden_dim = configs.hidden_dim
        self.max_char_len = configs.max_char_len
        self.char_dim = configs.char_dim
        self.char_lstm_dim = configs.char_lstm_dim
        self.chars_voab_num = dataManager.max_char_num

        self.pos_emb_dim = configs.pos_emb_dim

        self._filter_length_list = [1, 2, 3, 4, 5]
        self._nb_filter_list = [10, 10, 10, 10, 10]


        if configs.cell_type == 'LSTM':
            if self.biderectional:
                self.cell = tf.nn.rnn_cell.LSTMCell(self.hidden_dim)
            else:
                self.cell = tf.nn.rnn_cell.LSTMCell(2 * self.hidden_dim)
        else:
            if self.biderectional:
                self.cell = tf.nn.rnn_cell.GRUCell(self.hidden_dim)
            else:
                self.cell = tf.nn.rnn_cell.GRUCell(2 * self.hidden_dim)

        self.is_attention = configs.use_self_attention
        self.attention_dim = configs.attention_dim

        self.num_epochs = configs.epoch
        self.max_time_steps = configs.max_sequence_length

        self.num_tokens = dataManager.max_token_number
        self.num_classes = dataManager.max_label_number

        self.is_early_stop = configs.is_early_stop
        self.patient = configs.patient

        self.max_to_keep = configs.checkpoints_max_to_keep
        self.print_per_batch = configs.print_per_batch
        self.best_f1_val = 0

        self.reduce = -1.0


        if configs.optimizer == 'Adagrad':
            self.optimizer = tf.train.AdagradOptimizer(self.learning_rate)
        elif configs.optimizer == 'Adadelta':
            self.optimizer = tf.train.AdadeltaOptimizer(self.learning_rate)
        elif configs.optimizer == 'RMSprop':
            self.optimizer = tf.train.RMSPropOptimizer(self.learning_rate)
        elif configs.optimizer == 'GD':
            self.optimizer = tf.train.GradientDescentOptimizer(self.learning_rate)
        else:
            self.optimizer = tf.train.AdamOptimizer(self.learning_rate)

        self.initializer = tf.contrib.layers.xavier_initializer()
        self.global_step = tf.Variable(0, trainable=False, name="global_step", dtype=tf.int32)

        self.char_embedding = tf.get_variable("char_embedding", [self.chars_voab_num, self.char_dim],initializer=self.initializer, trainable=True)

        if configs.use_pretrained_embedding:
            embedding_matrix = dataManager.getEmbedding(configs.token_emb_dir)
            self.embedding = tf.Variable(embedding_matrix, trainable=False, name="emb", dtype=tf.float32)
        else:
            self.embedding = tf.get_variable("emb", [self.num_tokens, self.emb_dim], trainable=True,
                                             initializer=self.initializer)
        self.build()
        self.logger.info("model initialed...\n")

        self.sess = tf.Session(config=tf.ConfigProto(allow_soft_placement=True))


    def MultiConvolutional3D(self, input_data, filter_length_list, nb_filter_list, padding='VALID', pooling='max', name='Convolutional3D'):
        assert padding in ('VALID'), 'Unknow padding %s' % padding

        char_dim = int(input_data.get_shape()[-1])  
        input_data = tf.expand_dims(input_data, -1) 
        pooling_outpouts = []
        for i in range(len(filter_length_list)):
            filter_length = filter_length_list[i]
            nb_filter = nb_filter_list[i]
            with tf.variable_scope('%s_%d' % (name, filter_length)) as scope:

                conv_output = tf.contrib.layers.conv3d(
                    inputs=input_data,
                    num_outputs=nb_filter,
                    kernel_size=[1, filter_length, char_dim],
                    padding=padding)

                act_output = tf.nn.relu(conv_output)

                if pooling == 'max':
                    pooling_output = tf.reduce_max(tf.squeeze(act_output, [-2]), 2)
                elif pooling == 'mean':
                    pooling_output = tf.reduce_mean(tf.squeeze(act_output, [-2]), 2)
                else:
                    raise Exception('pooling must in (max, mean)!')
                pooling_outpouts.append(pooling_output)

                scope.reuse_variables()

        output = tf.concat(pooling_outpouts, axis=-1)
        return output

    def get_reduce_score(self,logits):
        inside = [0, 0, self.reduce, 0, 0, 0, 0, 0]
        reduce_score = []
        for l in range(self.batch_size*self.max_time_steps):
            reduce_score.append(inside)
        reduce_score = tf.convert_to_tensor(np.array(reduce_score).astype(np.float32))
        print(reduce_score.shape)
        return tf.add(logits,reduce_score)


    def build(self):
        self.inputs = tf.placeholder(tf.int32, [None, self.max_time_steps])
        self.inputs_char = tf.placeholder(tf.int32, [None, None, None])#[batch_size,max_seq_len,max_char_len]

        self.targets = tf.placeholder(tf.int32, [None, self.max_time_steps])

        char_emb = tf.nn.embedding_lookup(self.char_embedding,self.inputs_char)#[batch_size,max_seq_len,max_char_len,char_dim]
        self.inputs_char_emb = self.MultiConvolutional3D(char_emb, self._filter_length_list,self._nb_filter_list)
        print(self.inputs_char_emb.shape)
        self.inputs_glove_emb = tf.nn.embedding_lookup(self.embedding,self.inputs)
        self.inputs_emb = tf.concat([self.inputs_glove_emb,self.inputs_char_emb],axis=-1)
        print(self.inputs_emb.shape)
        self.epoch = tf.placeholder(tf.int32, [None])
        self.inputs_emb = tf.transpose(self.inputs_emb, [1, 0, 2])
        self.inputs_emb = tf.reshape(self.inputs_emb, [-1, self.char_lstm_dim+self.emb_dim])
        self.inputs_emb = tf.split(self.inputs_emb, self.max_time_steps, 0)
        if self.biderectional:
            lstm_cell_fw = self.cell
            lstm_cell_bw = self.cell
            if self.is_training:
                lstm_cell_fw = tf.nn.rnn_cell.DropoutWrapper(lstm_cell_fw, output_keep_prob=(1 - self.dropout_rate))
                lstm_cell_bw = tf.nn.rnn_cell.DropoutWrapper(lstm_cell_bw, output_keep_prob=(1 - self.dropout_rate))
            lstm_cell_fw = tf.nn.rnn_cell.MultiRNNCell([lstm_cell_fw] * self.num_layers)
            lstm_cell_bw = tf.nn.rnn_cell.MultiRNNCell([lstm_cell_bw] * self.num_layers)
            self.length = tf.reduce_sum(tf.sign(self.inputs), reduction_indices=1)
            self.length = tf.cast(self.length, tf.int32)
            outputs, _, _ = tf.contrib.rnn.static_bidirectional_rnn(
                lstm_cell_fw,
                lstm_cell_bw,
                self.inputs_emb,
                dtype=tf.float32,
                sequence_length=self.length
            )
        else:
            lstm_cell = self.cell
            if self.is_training:
                lstm_cell = tf.nn.rnn_cell.DropoutWrapper(lstm_cell, output_keep_prob=(1 - self.dropout_rate))
            lstm_cell = tf.nn.rnn_cell.MultiRNNCell([lstm_cell] * self.num_layers)
            self.length = tf.reduce_sum(tf.sign(self.inputs), reduction_indices=1)
            self.length = tf.cast(self.length, tf.int32)

            outputs, _ = tf.contrib.rnn.static_rnn(
                lstm_cell,
                self.inputs_emb,
                dtype=tf.float32,
                sequence_length=self.length
            )
        outputs = tf.concat(outputs, 1)
        outputs = tf.reshape(outputs, [self.batch_size, self.max_time_steps, self.hidden_dim * 2])
        if self.is_attention:
            H1 = tf.reshape(outputs, [-1, self.hidden_dim * 2])
            W_a1 = tf.get_variable("W_a1", shape=[self.hidden_dim * 2, self.attention_dim],
                                   initializer=self.initializer, trainable=True)
            u1 = tf.matmul(H1, W_a1)
            H2 = tf.reshape(tf.identity(outputs), [-1, self.hidden_dim * 2])
            W_a2 = tf.get_variable("W_a2", shape=[self.hidden_dim * 2, self.attention_dim],
                                   initializer=self.initializer, trainable=True)
            u2 = tf.matmul(H2, W_a2)
            u1 = tf.reshape(u1, [self.batch_size, self.max_time_steps, self.hidden_dim * 2])
            u2 = tf.reshape(u2, [self.batch_size, self.max_time_steps, self.hidden_dim * 2])
            u = tf.matmul(u1, u2, transpose_b=True)
            A = tf.nn.softmax(u, name="attention")
            outputs = tf.matmul(A, tf.reshape(tf.identity(outputs),
                                              [self.batch_size, self.max_time_steps, self.hidden_dim * 2]))


        self.outputs = tf.reshape(outputs, [-1, self.hidden_dim * 2])
        self.softmax_w = tf.get_variable("softmax_w", [self.hidden_dim * 2, self.num_classes],
                                         initializer=self.initializer)
        self.softmax_b = tf.get_variable("softmax_b", [self.num_classes], initializer=self.initializer)
        self.logits = tf.matmul(self.outputs, self.softmax_w) + self.softmax_b

        self.logits = self.get_reduce_score(self.logits)

        self.logits = tf.reshape(self.logits, [self.batch_size, self.max_time_steps, self.num_classes])
        # print(self.logits.get_shape().as_list())
        if not self.is_crf:
            # softmax
            softmax_out = tf.nn.softmax(self.logits, axis=-1)

            self.batch_pred_sequence = tf.cast(tf.argmax(softmax_out, -1), tf.int32)
            losses = tf.nn.sparse_softmax_cross_entropy_with_logits(logits=self.logits, labels=self.targets)
            mask = tf.sequence_mask(self.length)

            self.losses = tf.boolean_mask(losses, mask)

            self.loss = tf.reduce_mean(losses)
        else:
            # crf
            self.log_likelihood, self.transition_params = tf.contrib.crf.crf_log_likelihood(
                self.logits, self.targets, self.length)
            self.batch_pred_sequence, self.batch_viterbi_score = tf.contrib.crf.crf_decode(self.logits,
                                                                                           self.transition_params,
                                                                                           self.length)
            alpha = tf.constant(8,dtype = tf.int32) 
            tau = tf.constant(0.5, dtype = tf.float64) 
            self.w = tf.cond(self.epoch[0]>=45 and self.epoch[0]<=75 , lambda:1, lambda:tf.cond(self.epoch[0]>75 , lambda:1/(1+tf.exp(-alpha*(-self.log_likelihood-tau))), lambda:1/(1+tf.exp(alpha*(-self.log_likelihood-tau)))))
            self.new_loss_everysample = tf.multiply(self.w,-self.log_likelihood)
            self.loss = tf.reduce_mean(self.new_loss_everysample)

        self.train_summary = tf.summary.scalar("loss", self.loss)
        self.dev_summary = tf.summary.scalar("loss", self.loss)

        self.opt_op = self.optimizer.minimize(self.loss, global_step=self.global_step)

    def train(self):
        X_train, y_train,X_train_char, X_val, y_val,X_val_char = self.dataManager.getTrainingSet()
        tf.initialize_all_variables().run(session=self.sess)

        saver = tf.train.Saver(max_to_keep=self.max_to_keep)
        tf.summary.merge_all()
        train_writer = tf.summary.FileWriter(self.logdir + "/training_loss", self.sess.graph)
        dev_writer = tf.summary.FileWriter(self.logdir + "/validating_loss", self.sess.graph)

        num_iterations = int(math.ceil(1.0 * len(X_train) / self.batch_size))
        num_val_iterations = int(math.ceil(1.0 * len(X_val) / self.batch_size))

        cnt = 0
        cnt_dev = 0
        unprogressed = 0
        very_start_time = time.time()
        best_at_epoch = 0
        self.logger.info("\ntraining starting" + ("+" * 20))
        for epoch in range(self.num_epochs):
            start_time = time.time()
            # shuffle train at each epoch
            sh_index = np.arange(len(X_train))
            np.random.shuffle(sh_index)
            X_train = X_train[sh_index]
            y_train = y_train[sh_index]
            X_train_char = X_train_char[sh_index]

            self.logger.info("\ncurrent epoch: %d" % (epoch))
            for iteration in range(num_iterations):
                X_train_batch, y_train_batch,X_train_char_batch = self.dataManager.nextBatch(X_train, y_train,X_train_char,
                                                                          start_index=iteration * self.batch_size)
                _, loss_train, train_batch_viterbi_sequence, train_summary, loss_one= \
                    self.sess.run([
                        self.opt_op,
                        self.loss,
                        self.batch_pred_sequence,
                        self.train_summary,
                        self.log_likelihood
                    ],
                        feed_dict={
                            self.epoch: epoch,
                            self.inputs: X_train_batch,
                            self.inputs_char:X_train_char_batch,
                            self.targets: y_train_batch,
                        })

                if iteration % self.print_per_batch == 0:
                    cnt += 1
                    train_writer.add_summary(train_summary, cnt)

                    measures = metrics(y_train_batch,
                                       train_batch_viterbi_sequence,
                                       self.measuring_metrics, self.dataManager)

                    res_str = ''
                    for k, v in measures.items():
                        res_str += (k + ": %.4f " % v)
                    self.logger.info("training batch: %5d, loss: %.5f, %s" % (iteration, loss_train, res_str))
                    self.logger.info(loss_train, loss_one)

            loss_vals = list()

            y_val_pred = []
            y_val_true = []

            for iterr in range(num_val_iterations):
                cnt_dev += 1
                X_val_batch, y_val_batch,X_val_char_batch = self.dataManager.nextBatch(X_val, y_val,X_val_char,start_index=iterr * self.batch_size)
                loss_val, val_batch_viterbi_sequence, dev_summary = \
                    self.sess.run([
                        self.loss,
                        self.batch_pred_sequence,
                        self.dev_summary
                    ],
                        feed_dict={
                            self.inputs: X_val_batch,
                            self.inputs_char:X_val_char_batch,
                            self.targets: y_val_batch,
                        })
                for j in range(len(y_val_batch)):
                    y_val_pred.append(val_batch_viterbi_sequence[j])
                    y_val_true.append(y_val_batch[j])
                dev_writer.add_summary(dev_summary, cnt_dev)
                loss_vals.append(loss_val)
            y_val_true = y_val_true[0:len(X_val)]
            y_val_pred = y_val_pred[0:len(X_val)]
            measures = metrics(y_val_true, y_val_pred,
                               self.measuring_metrics, self.dataManager)

            time_span = (time.time() - start_time) / 60
            val_res_str = ''
            dev_f1_avg = 0
            for k, v in measures.items():
                val_res_str += (k + ": %.3f " % v)
                if k == 'f1': dev_f1_avg = measures[k]


            self.logger.info("time consumption:%.2f(min),  validation loss: %.5f, %s" %
                             (time_span, np.array(loss_vals).mean(), val_res_str))

            if dev_f1_avg > self.best_f1_val:
                unprogressed = 0
                self.best_f1_val = dev_f1_avg
                best_at_epoch = epoch
                saver.save(self.sess, self.checkpoints_dir + "/" + self.checkpoint_name, global_step=self.global_step)
                self.logger.info("saved the new best model with f1: %.3f" % (self.best_f1_val))
            else:
                unprogressed += 1

            if self.is_early_stop:
                if unprogressed >= self.patient:
                    self.logger.info("early stopped, no progress obtained within %d epochs" % self.patient)
                    self.logger.info("overall best f1 is %f at %d epoch" % (self.best_f1_val, best_at_epoch))
                    self.logger.info(
                        "total training time consumption: %.3f(min)" % ((time.time() - very_start_time) / 60))
                    self.sess.close()
                    return
        self.logger.info("overall best f1 is %f at %d epoch" % (self.best_f1_val, best_at_epoch))
        self.logger.info("total training time consumption: %.3f(min)" % ((time.time() - very_start_time) / 60))
        self.sess.close()

    def test(self):
        start_time = time.time()
        if self.is_real_test:
            X_test,X_test_pos, y_test_psyduo_label, X_test_str, X_test_char,rule_metrics_entities = self.dataManager.getTestingSet()
        else:
            X_test, y_test_psyduo_label, X_test_str, X_test_char = self.dataManager.getTestingSet()



        num_iterations = int(math.ceil(1.0 * len(X_test) / self.batch_size))
        self.logger.info("total number of testing iterations: " + str(num_iterations))


        self.logger.info("loading model parameter\n")
        tf.initialize_all_variables().run(session=self.sess)

        saver = tf.train.Saver()
        saver.restore(self.sess, tf.train.latest_checkpoint(self.checkpoints_dir))

        tokens = []
        labels = []
        real_labels = []
        real_entities = []
        real_entities_types = []
        entities = []
        entities_pos = []
        entities_types = []
        y_test_true = []
        y_test_pred = []
        self.logger.info("\ntesting starting" + ("+" * 20))
        for i in range(num_iterations):
            self.logger.info("batch: " + str(i + 1))
            X_test_batch,X_test_pos_batch,y_test_psyduo_label_batch,X_test_str_batch,X_test_char_batch,is_less_batch \
                = self.dataManager.nextTestBatch(X_test,X_test_pos, y_test_psyduo_label, X_test_str, X_test_char,start_index=i*self.batch_size)

            predicts_label_id, results, token, real_entity, entity,entity_pos, real_entities_type, entities_type, _, _ = self.predictBatch(
                self.sess, X_test_batch,X_test_pos_batch,y_test_psyduo_label_batch,X_test_str_batch,X_test_char_batch)

            for j in range(len(y_test_psyduo_label_batch)):
                y_test_true.append(y_test_psyduo_label_batch[j].tolist())
                y_test_pred.append(predicts_label_id[j].tolist())
            labels.extend(results)
            tokens.extend(token)
            real_entities.extend(real_entity)
            entities.extend(entity)
            entities_pos.extend(entity_pos)
            real_entities_types.extend(real_entities_type)
            entities_types.extend(entities_type)
        labels = labels[0:len(X_test)]
        tokens = tokens[0:len(X_test)]
        real_entities = real_entities[0:len(X_test)]
        entities = entities[0:len(X_test)]
        entities_pos = entities_pos[0:len(X_test)]
        real_entities_types = real_entities_types[0:len(X_test)]
        entities_types = entities_types[0:len(X_test)]
        y_test_true = y_test_true[0:len(X_test)]
        y_test_pred = y_test_pred[0:len(X_test)]
        measures = metrics(y_test_true, y_test_pred,
                           self.measuring_metrics, self.dataManager)


        if self.is_real_test:
            raw_real = xlrd.open_workbook(self.raw_real_sens_file, 'rb').sheet_by_name(u'Sheet1')
            raw_real_ids = raw_real.col_values(1)
            raw_real_titles = raw_real.col_values(2)
            raw_real_meetings = raw_real.col_values(3)
            raw_real_years = raw_real.col_values(4)
            raw_real_paras = raw_real.col_values(5)
            raw_real_sens = raw_real.col_values(6)

            text_metrics = xlrd.open_workbook('black_Metrics_list.xls', 'rb').sheet_by_name(u'sheet1')
            black_metrics = text_metrics.col_values(0)
            text_methods = xlrd.open_workbook('black_Methods_list.xls', 'rb').sheet_by_name(u'sheet1')
            black_methods = text_methods.col_values(0)
            text_datasets = xlrd.open_workbook('black_Datasets_list.xls', 'rb').sheet_by_name(u'sheet1')
            black_datasets = text_datasets.col_values(0)

            def password_check_contain_upper(password):
                pattern = re.compile('[A-Z]+')
                match = pattern.findall(password)
                if match:
                    return True
                else:
                    return False
            def get_entitiesAndsentences(entities, entities_types,entities_pos,tokens):
                data_entites_list = []
                method_entitys_list = []
                index_entitys_list = []
                delet_pos_dic = ['VB', 'VBP', 'VBZ', 'FW','DT','PRP','PRP$']
                delet_dic = ['algorithm','approach','benchmark','classifier','corpora','data','database','dataset','method','model','network',
                             'algorithms','approaches','benchmarks','classifiers','corporas','corpus','datas','databases','datasets','methods','models','networks']
                for row in range(len(entities_types)):
                    data_entity = []
                    method_entity = []
                    index_entity = []
                    token_lower = ' '.join(tokens[row]).lower()
                    for num in range(len(entities_types[row])):
                        if entities_types[row][num] == 'M':
                            entity_str = []
                            str_list = re.split(' ', entities[row][num])
                            str_pos_list = re.split(' ', entities_pos[row][num])
                            if len(entities[row][num]) > 1 and ('NN'in entities_pos[row][num] or 'JJ' in entities_pos[row][num]) \
                                    and password_check_contain_upper(entities[row][num]) and entities[row][num].replace('-','').lower() not in delet_dic and entities[row][num].replace('-','').lower() not in black_methods:
                                for i in range(len(str_list)):
                                    if str_pos_list[i] not in delet_pos_dic and str_list[i].replace('-','') not in delet_dic and str_list[i].replace('-','').lower() not in black_methods:
                                        entity_str.append(str_list[i])
                            if entity_str != []:
                                entity_str = ' '.join(entity_str)
                                if len(entity_str) > 1 and password_check_contain_upper(entity_str):
                                    if '/' in entity_str:
                                        entity_str_split = re.split('/', entity_str)
                                        for val in entity_str_split:
                                            if len(val) > 1 and password_check_contain_upper(val):
                                                method_entity.append(val)
                                    else:
                                        method_entity.append(entity_str)

                        elif entities_types[row][num] == 'D':
                            entity_str = []
                            if 'benchmark' in token_lower or 'corpora' in token_lower or 'data' in token_lower or 'corpus'in token_lower:
                                str_list = re.split(' ', entities[row][num])
                                str_pos_list = re.split(' ', entities_pos[row][num])
                                if len(entities[row][num]) > 1 and (
                                        'NN' in entities_pos[row][num] or 'JJ' in entities_pos[row][num]) \
                                        and entities[row][num].replace('-', '').lower() not in delet_dic and \
                                        entities[row][num].replace('-', '').lower() not in black_datasets:
                                    for i in range(len(str_list)):
                                        if str_pos_list[i] not in delet_pos_dic and str_list[i].replace('-',
                                                                                                        '') not in delet_dic and \
                                                str_list[i].replace('-', '').lower() not in black_datasets:
                                            entity_str.append(str_list[i])

                            if entity_str != []:
                                entity_str = ' '.join(entity_str)
                                if len(entity_str) > 1:
                                    if '/' in entity_str:
                                        entity_str_split = re.split('/', entity_str)
                                        for val in entity_str_split:
                                            if len(val) > 1:
                                                data_entity.append(val)
                                    else:
                                        data_entity.append(entity_str)

                        else:
                            entity_str = []
                            str_list = re.split(' ',entities[row][num])
                            str_pos_list = re.split(' ', entities_pos[row][num])
                            if len(entities[row][num]) > 1 and ('NN'in entities_pos[row][num] or 'JJ' in entities_pos[row][num]) \
                                    and entities[row][num].replace('-','').lower() not in delet_dic and entities[row][num].replace('-','').lower() not in black_metrics:
                                for i in range(len(str_list)):
                                    if str_pos_list[i] not in delet_pos_dic and str_list[i].replace('-','') not in delet_dic and str_list[i].replace('-','').lower() not in black_metrics:
                                        entity_str.append(str_list[i])
                            if entity_str != []:
                                entity_str = ' '.join(entity_str)
                                if len(entity_str) > 1:
                                    if '/' in entity_str:
                                        entity_str_split = re.split('/', entity_str)
                                        for val in entity_str_split:
                                            if len(val) > 1:
                                                index_entity.append(val)
                                    else:
                                        index_entity.append(entity_str)
                    index_entity += rule_metrics_entities[row]

                    data_entites_list.append(data_entity)
                    method_entitys_list.append(method_entity)
                    index_entitys_list.append(index_entity)
                data_outputs = []
                method_outputs = []
                index_outputs = []
                for row in range(len(entities)):
                    data_output = ';'.join(data_entites_list[row])
                    method_output = ';'.join(method_entitys_list[row])
                    index_ouput = ';'.join(index_entitys_list[row])
                    data_outputs.append(data_output)
                    method_outputs.append(method_output)
                    index_outputs.append(index_ouput)
                return method_outputs, data_outputs, index_outputs


            def get_conj_entities(sentence_tree, sentence, sentence_pos):
                del sentence_tree[0]
                length = 0
                for tup in sentence_tree:
                    if tup[0] != 'ROOT':
                        length += 1
                    else:
                        break
                sentence_tree = sentence_tree[0:length]
                conj_nodes = []
                for tup in sentence_tree:
                    if tup[0] == 'cc':
                        conj_node = []
                        node = tup[1]
                        conj_node.append(node)
                        for in_tup in sentence_tree:
                            if in_tup[1] == node and in_tup[0] == 'conj':
                                conj_node.append(in_tup[2])
                        conj_nodes.append(conj_node)
                entities = []
                entities_pos = []
                for l_val in conj_nodes:
                    conj_entity = []
                    conj_pos_entity = []
                    for val in l_val:
                        entity = []
                        entity.append(val)
                        for tup in sentence_tree:
                            if (tup[0] == 'compound' or tup[0] == 'amod') and tup[1] == val:
                                entity.append(tup[2])
                        entity.sort()
                        entity_str = ' '.join(sentence[val - 1] for val in entity)
                        entity_pos = ' '.join(sentence_pos[val - 1] for val in entity)
                        conj_entity.append(entity_str)
                        conj_pos_entity.append(entity_pos)
                    entities.append(conj_entity)
                    entities_pos.append(conj_pos_entity)
                return entities, entities_pos

            def get_entitiesFromGRN(real_sens,entities, entities_types,entities_pos):
                nlp = StanfordCoreNLP('stanford-corenlp-full-2018-10-05')
                new_entities = []
                new_entities_types = []
                new_entities_pos = []
                for row in range(len(real_sens)):
                    if ((' and ' in real_sens[row]) or (' or ' in real_sens[row]) ) and (entities[row] != []):
                        sentence = real_sens[row].replace('“', '"').replace('”', '"').replace('’', '\'').replace('‘',
                                                                                                                  '\'').replace(
                            '！', '!').replace('\t', ' ')
                        sentence = re.sub("\[.*[0-9]{4}\s*\]", " ", sentence)
                        pos_sentence = nlp.pos_tag(sentence)
                        sentence_tree = nlp.dependency_parse(sentence)
                        sentence = [tp[0] for tp in pos_sentence]
                        sentence_pos = [tp[1] for tp in pos_sentence]
                        print(row)
                        grn_entities, grn_entities_pos = get_conj_entities(sentence_tree,sentence,sentence_pos)
                        new_entity = []
                        new_entity_pos = []
                        new_entity_type = []
                        delet_id = []
                        for i in range(len(grn_entities)):
                            judge = False
                            type_str = ''
                            for j in range(len(grn_entities[i])):
                                for num in range(len(entities[row])):
                                    if (grn_entities[i][j] in entities[row][num] ) or (entities[row][num] in grn_entities[i][j]):
                                        delet_id.append(num)
                                        judge = True
                                        type_str = entities_types[row][num]

                            if judge:
                                new_entity+=grn_entities[i]
                                new_entity_pos+=grn_entities_pos[i]
                                new_entity_type+=[type_str]*len(grn_entities[i])

                        delet_id_dic = []
                        for l in range(len(delet_id)):
                            if delet_id[l] not in delet_id_dic:
                                delet_id_dic.append(delet_id[l])

                        delet_id_dic.sort()

                        count = 0
                        for l in delet_id_dic:
                            del entities[row][l-count]
                            del entities_pos[row][l-count]
                            del entities_types[row][l-count]
                            count+=1

                        new_entity+=entities[row]
                        new_entity_pos+=entities_pos[row]
                        new_entity_type+=entities_types[row]

                        new_entities.append(new_entity)
                        new_entities_types.append(new_entity_type)
                        new_entities_pos.append(new_entity_pos)
                    else:
                        new_entities.append(entities[row])
                        new_entities_types.append(entities_types[row])
                        new_entities_pos.append(entities_pos[row])
                nlp.close()
                return new_entities,new_entities_types,new_entities_pos

            methods_list, datasets_list, indexes_list = get_entitiesAndsentences(entities, entities_types, entities_pos,tokens)
            new_entities, new_entities_types, new_entities_pos = get_entitiesFromGRN(raw_real_sens,entities,entities_types,entities_pos)
            new_methods_list, new_datasets_list, new_indexes_list = get_entitiesAndsentences(new_entities, new_entities_types, new_entities_pos,tokens)


            def test_output(real_ids, real_titles, real_meetings, real_years, real_paras, real_sens, methods_list,
                            datasets_list, indexes_list,new_methods_list, new_datasets_list, new_indexes_list, output_file):

                xls = openpyxl.Workbook()
                sheet1 = xls.get_sheet_by_name('Sheet')
                for i in range(len(real_ids)):
                    sheet1.cell(row=i+1, column=1, value=real_ids[i])
                    sheet1.cell(row=i+1, column=2, value=real_titles[i])
                    sheet1.cell(row=i+1, column=3, value=real_meetings[i])
                    sheet1.cell(row=i+1, column=4, value=real_years[i])
                    sheet1.cell(row=i+1, column=5, value=real_paras[i])
                    sheet1.cell(row=i+1, column=6, value=real_sens[i])
                    sheet1.cell(row=i+1, column=7, value=methods_list[i])
                    sheet1.cell(row=i+1, column=8, value=datasets_list[i])
                    sheet1.cell(row=i+1, column=9, value=indexes_list[i])
                    sheet1.cell(row=i+1, column=10, value=new_methods_list[i])
                    sheet1.cell(row=i+1, column=11, value=new_datasets_list[i])
                    sheet1.cell(row=i+1, column=12, value=new_indexes_list[i])
                xls.save(output_file)

            test_output(raw_real_ids, raw_real_titles, raw_real_meetings, raw_real_years, raw_real_paras, raw_real_sens,
                        methods_list, datasets_list, indexes_list,new_methods_list, new_datasets_list, new_indexes_list , self.output_excel)
            self.logger.info("finished output excel for sentences!\n")

            if os.path.isfile(self.output_experiment_file) and os.path.isfile(self.output_methods_file):
                Sentence_to_Paper(self.output_methods_file,self.output_experiment_file, self.papaer_unit_excel_file)
            self.logger.info("finished output excel for Papers!\n")


            time_span = (time.time() - start_time) / 60
            self.logger.info("time consumption:%.2f(min)" % (time_span))

            self.sess.close()
            tf.reset_default_graph()

        else:
            self.raw_sentences = xlrd.open_workbook('test.xls', 'rb').sheet_by_name(u'sheet1')
            self.raw_sens = self.raw_sentences.col_values(1)
            self.raw_ids = self.raw_sentences.col_values(0)
            self.raw_class = self.raw_sentences.col_values(2)
            time_span = (time.time() - start_time) / 60
            test_res_str = ''
            for k, v in measures.items():
                test_res_str += (k + ": %.3f " % v)
            self.logger.info("time consumption:%.2f(min), %s" % (time_span, test_res_str))
            id_array = []
            class_array = []
            new_sentences = []
            new_real_entities = []
            new_entities = []
            count = 0
            for l in range(len(real_entities)):
                real_entitie = real_entities[l]
                real_entities_type = real_entities_types[l]
                entitie = entities[l]
                entities_type = entities_types[l]
                if real_entitie == entitie and real_entities_type == entities_type:
                    pass
                else:
                    count += 1
                    id_array.append(self.raw_ids[l])
                    new_sentences.append(self.raw_sens[l])
                    class_array.append(self.raw_class[l])
                    new_real_list = [real_entitie[i] + " " + "(" + real_entities_type[i] + ")" for i in range(len(real_entitie))]
                    new_pre_list = [entitie[i] + " " + "(" + entities_type[i] + ")" for i in range(len(entitie))]
                    new_real_str = ';'.join(new_real_list)
                    new_pre_str = ';'.join(new_pre_list)
                    new_real_entities.append(new_real_str)
                    new_entities.append(new_pre_str)
            f = xlwt.Workbook()
            sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)
            for i in range(len(id_array)):
                sheet1.write(i, 0, id_array[i])
                sheet1.write(i, 1, new_sentences[i])
                sheet1.write(i, 2, new_real_entities[i])
                sheet1.write(i, 3, new_entities[i])
                sheet1.write(i, 4, class_array[i])
            f.save(str(self.reduce)+'output.xls')

            print("the num of wrong sentences is:",count)
            self.sess.close()

    def predictBatch(self, sess, X,X_pos, y_psydo_label, X_test_str_batch,X_test_char):
        entity_list = []
        tokens = []
        predicts_labels_entitylevel = []
        indexs = []
        predicts_labels_tokenlevel = []
        real_entity_list = []
        entity_pos_list = []
        real_labels_entitylevel = []
        real_indexs = []

        predicts_label_id, lengths = \
            sess.run([
                self.batch_pred_sequence,
                self.length
            ],
                feed_dict={
                    self.inputs: X,
                    self.inputs_char:X_test_char,
                    self.targets: y_psydo_label,
                })

        for i in range(len(lengths)):
            x_ = [val for val in X_test_str_batch[i, 0:lengths[i]]]
            tokens.append(x_)
            x_pos = [val for val in X_pos[i, 0:lengths[i]]]

            y_pred = [str(self.dataManager.id2label[val]) for val in predicts_label_id[i, 0:lengths[i]]]
            predicts_labels_tokenlevel.append(y_pred)

            y_ = [str(self.dataManager.id2label[val]) for val in y_psydo_label[i, 0:lengths[i]]]

            entitys, entity_labels, labled_indexs , entitys_pos = extractEntity(x_,x_pos, y_pred, self.dataManager)
            real_entitys, real_entity_labels, real_labled_indexs,_ = extractEntity(x_,x_pos, y_, self.dataManager)
            real_entity_list.append(real_entitys)
            entity_pos_list.append(entitys_pos)
            entity_list.append(entitys)
            real_labels_entitylevel.append(real_entity_labels)
            predicts_labels_entitylevel.append(entity_labels)
            real_indexs.append(real_labled_indexs)
            indexs.append(labled_indexs)

        return predicts_label_id, predicts_labels_tokenlevel, tokens, real_entity_list, entity_list,entity_pos_list, real_labels_entitylevel, predicts_labels_entitylevel, real_indexs, indexs

    def soft_load(self):
        self.logger.info("loading model parameter")
        tf.initialize_all_variables().run(session=self.sess)
        saver = tf.train.Saver()
        saver.restore(self.sess, tf.train.latest_checkpoint(self.checkpoints_dir))
        self.logger.info("loading model successfully")